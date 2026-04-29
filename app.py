import streamlit as st
import cv2
import face_recognition
import numpy as np
import sqlite3
import pickle
import pandas as pd
import os
import time
from datetime import datetime
from io import BytesIO
import openpyxl

# ─── CONFIG ────────────────────────────────────────────────────────────────────
DB_PATH = "attendance.db"
ENCODINGS_FILE = "face_encodings.pkl"

st.set_page_config(
    page_title="FaceAttend AI",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;700;800&display=swap');

* { font-family: 'Syne', sans-serif; }
code, .stCode { font-family: 'Space Mono', monospace !important; }

/* Main background */
.stApp {
    background: #0a0a0f;
    color: #e8e8f0;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: #0f0f1a !important;
    border-right: 1px solid #1e1e3a;
}
[data-testid="stSidebar"] * { color: #c8c8e0 !important; }
[data-testid="stSidebar"] .stRadio label {
    font-size: 14px !important;
    padding: 8px 0 !important;
    letter-spacing: 0.05em !important;
}

/* Headers */
h1, h2, h3 { font-family: 'Syne', sans-serif !important; font-weight: 800 !important; }
h1 { 
    font-size: 2.2rem !important; 
    background: linear-gradient(135deg, #7c5cbf, #4f8ef7, #00d4aa);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: -0.02em;
}

/* Metric cards */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, #1a1a2e, #16213e);
    border: 1px solid #2a2a4a;
    border-radius: 12px;
    padding: 20px !important;
    box-shadow: 0 4px 20px rgba(79, 142, 247, 0.08);
}
[data-testid="stMetricValue"] {
    font-size: 2.5rem !important;
    font-weight: 800 !important;
    color: #4f8ef7 !important;
}
[data-testid="stMetricLabel"] { color: #888 !important; font-size: 12px !important; letter-spacing: 0.1em; }

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, #4f8ef7, #7c5cbf) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: 0.05em !important;
    padding: 10px 24px !important;
    transition: all 0.2s ease !important;
    font-family: 'Syne', sans-serif !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(79, 142, 247, 0.4) !important;
}

/* Danger button */
button[kind="secondary"] {
    background: linear-gradient(135deg, #c0392b, #e74c3c) !important;
}

/* Input fields */
.stTextInput input, .stSelectbox select {
    background: #1a1a2e !important;
    border: 1px solid #2a2a4a !important;
    color: #e8e8f0 !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
}
.stTextInput input:focus {
    border-color: #4f8ef7 !important;
    box-shadow: 0 0 0 2px rgba(79, 142, 247, 0.2) !important;
}

/* Tables */
.stDataFrame {
    border: 1px solid #2a2a4a !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}

/* Success / Warning / Error */
.stSuccess { background: rgba(0, 212, 170, 0.1) !important; border-left: 3px solid #00d4aa !important; }
.stWarning { background: rgba(255, 180, 50, 0.1) !important; border-left: 3px solid #ffb432 !important; }
.stError { background: rgba(255, 80, 80, 0.1) !important; border-left: 3px solid #ff5050 !important; }
.stInfo { background: rgba(79, 142, 247, 0.1) !important; border-left: 3px solid #4f8ef7 !important; }

/* Section divider */
hr { border-color: #1e1e3a !important; }

/* Card containers */
.card {
    background: linear-gradient(135deg, #1a1a2e, #16213e);
    border: 1px solid #2a2a4a;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 16px;
}

/* Selectbox */
[data-testid="stSelectbox"] > div > div {
    background: #1a1a2e !important;
    border-color: #2a2a4a !important;
    color: #e8e8f0 !important;
}
</style>
""", unsafe_allow_html=True)

# ─── DATABASE ──────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            roll TEXT UNIQUE NOT NULL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            roll TEXT,
            subject TEXT,
            date TEXT,
            time TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_name TEXT UNIQUE NOT NULL
        )
    """)
    conn.commit()
    conn.close()

# ─── ENCODING HELPERS ──────────────────────────────────────────────────────────
def load_encodings():
    if os.path.exists(ENCODINGS_FILE):
        with open(ENCODINGS_FILE, "rb") as f:
            return pickle.load(f)
    return {"encodings": [], "names": [], "rolls": []}

def save_encodings(data):
    with open(ENCODINGS_FILE, "wb") as f:
        pickle.dump(data, f)

# ─── NAVIGATION ────────────────────────────────────────────────────────────────
init_db()

st.sidebar.markdown("## 🎯 **FaceAttend AI**")
st.sidebar.markdown("---")
page = st.sidebar.radio("Navigate", [
    "📊 Dashboard",
    "👤 Register Student",
    "📚 Subject Management",
    "📷 Mark Attendance",
    "📋 Reports",
    "📤 Export Excel",
    "🔄 Reset System"
])

# ══════════════════════════════════════════════════════════════════════════════
# 1. DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.title("📊 Dashboard")
    st.markdown("Real-time system statistics")
    st.markdown("---")

    conn = sqlite3.connect(DB_PATH)
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_records  = conn.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
    total_subjects = conn.execute("SELECT COUNT(*) FROM subjects").fetchone()[0]
    conn.close()

    c1, c2, c3 = st.columns(3)
    c1.metric("👤 Total Students", total_students)
    c2.metric("✅ Attendance Records", total_records)
    c3.metric("📚 Subjects", total_subjects)

    st.markdown("---")

    # Recent attendance
    conn = sqlite3.connect(DB_PATH)
    recent = pd.read_sql_query(
        "SELECT name, roll, subject, date, time FROM attendance ORDER BY id DESC LIMIT 10", conn
    )
    conn.close()

    if not recent.empty:
        st.markdown("### 🕐 Recent Attendance")
        st.dataframe(recent, use_container_width=True)
    else:
        st.info("No attendance records yet. Start marking attendance to see data here.")

# ══════════════════════════════════════════════════════════════════════════════
# 2. REGISTER STUDENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "👤 Register Student":
    st.title("👤 Register Student")
    st.markdown("Capture 20 face samples for facial recognition")
    st.markdown("---")

    # ── Delete Student ────────────────────────────────────────────────────────
    with st.expander("🗑️ Delete a Student", expanded=False):
        conn = sqlite3.connect(DB_PATH)
        all_students = pd.read_sql_query("SELECT name, roll FROM students ORDER BY name", conn)
        conn.close()

        if all_students.empty:
            st.info("No students registered yet.")
        else:
            del_options = [f"{r['name']} ({r['roll']})" for _, r in all_students.iterrows()]
            del_choice = st.selectbox("Select student to delete", del_options, key="del_student_select")
            del_roll = all_students.iloc[del_options.index(del_choice)]["roll"]

            if st.button("🗑️ Delete Student", key="del_student_btn"):
                # Remove from DB
                conn = sqlite3.connect(DB_PATH)
                conn.execute("DELETE FROM students WHERE roll=?", (del_roll,))
                conn.execute("DELETE FROM attendance WHERE roll=?", (del_roll,))
                conn.commit()
                conn.close()

                # Remove from encodings file
                enc_data = load_encodings()
                indices_to_keep = [i for i, r in enumerate(enc_data["rolls"]) if r != del_roll]
                enc_data["encodings"] = [enc_data["encodings"][i] for i in indices_to_keep]
                enc_data["names"]     = [enc_data["names"][i]     for i in indices_to_keep]
                enc_data["rolls"]     = [enc_data["rolls"][i]     for i in indices_to_keep]
                save_encodings(enc_data)

                st.success(f"✅ Student **{del_choice}** and all their attendance records deleted.")
                st.rerun()

    st.markdown("---")
    col1, col2 = st.columns([1, 1])

    with col1:
        name = st.text_input("Student Name", placeholder="e.g. Rahul Sharma")
        roll = st.text_input("Roll Number", placeholder="e.g. CS2024001")

    reg_btn = col1.button("🚀 Start Registration")

    if reg_btn:
        if not name.strip() or not roll.strip():
            st.error("Please enter both name and roll number.")
        else:
            name = name.strip()
            roll = roll.strip()

            # Check duplicate roll
            conn = sqlite3.connect(DB_PATH)
            exists = conn.execute("SELECT 1 FROM students WHERE roll=?", (roll,)).fetchone()
            conn.close()
            if exists:
                st.error(f"Roll number **{roll}** is already registered!")
            else:
                # Check face duplicate
                enc_data = load_encodings()

                frame_placeholder = col2.empty()
                status_placeholder = col1.empty()

                cap = cv2.VideoCapture(0)
                if not cap.isOpened():
                    st.error("Cannot open webcam. Please check camera connection.")
                else:
                    samples = []
                    count = 0
                    target = 20
                    start_time = time.time()

                    while count < target:
                        ret, frame = cap.read()
                        if not ret:
                            break

                        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        locs = face_recognition.face_locations(rgb)
                        encs = face_recognition.face_encodings(rgb, locs)

                        # Flip background first, then draw boxes on mirrored coords
                        display = cv2.flip(rgb, 1)
                        w = display.shape[1]

                        for (top, right, bottom, left), enc in zip(locs, encs):
                            # Check against existing encodings
                            if enc_data["encodings"]:
                                matches = face_recognition.compare_faces(enc_data["encodings"], enc, tolerance=0.5)
                                if True in matches:
                                    matched_idx = matches.index(True)
                                    matched_name = enc_data["names"][matched_idx]
                                    cap.release()
                                    st.error(f"⚠️ Face already registered as **{matched_name}**!")
                                    st.stop()

                            samples.append(enc)
                            count += 1

                            # Mirror the x coordinates for drawing
                            m_left  = w - right
                            m_right = w - left
                            cv2.rectangle(display, (m_left, top), (m_right, bottom), (0, 255, 0), 2)
                            cv2.putText(display, f"Capturing {count}/{target}", (m_left, top - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                        frame_placeholder.image(display, channels="RGB", use_column_width=True)
                        status_placeholder.info(f"📸 Capturing face samples: **{count}/{target}**")

                        if time.time() - start_time > 30:
                            break

                    cap.release()

                    if count >= target:
                        # Save to DB
                        conn = sqlite3.connect(DB_PATH)
                        conn.execute("INSERT INTO students (name, roll) VALUES (?, ?)", (name, roll))
                        conn.commit()
                        conn.close()

                        # Save encodings
                        for enc in samples:
                            enc_data["encodings"].append(enc)
                            enc_data["names"].append(name)
                            enc_data["rolls"].append(roll)
                        save_encodings(enc_data)

                        frame_placeholder.empty()
                        status_placeholder.empty()
                        st.success(f"✅ **Registration Complete!** {name} ({roll}) registered successfully with {count} face samples.")
                    else:
                        st.warning(f"⚠️ Only captured {count} samples. Please ensure proper lighting and face visibility.")

# ══════════════════════════════════════════════════════════════════════════════
# 3. SUBJECT MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📚 Subject Management":
    st.title("📚 Subject Management")
    st.markdown("Add and manage subjects for attendance tracking")
    st.markdown("---")

    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("### ➕ Add New Subject")
        subj_name = st.text_input("Subject Name", placeholder="e.g. Mathematics")
        if st.button("Add Subject"):
            if not subj_name.strip():
                st.error("Subject name cannot be empty.")
            else:
                normalized = subj_name.strip().lower()
                conn = sqlite3.connect(DB_PATH)
                exists = conn.execute("SELECT 1 FROM subjects WHERE subject_name=?", (normalized,)).fetchone()
                if exists:
                    st.error(f"Subject **{subj_name}** already exists!")
                else:
                    conn.execute("INSERT INTO subjects (subject_name) VALUES (?)", (normalized,))
                    conn.commit()
                    st.success(f"✅ Subject **{normalized}** added!")
                conn.close()

    with col2:
        st.markdown("### 📋 Existing Subjects")
        conn = sqlite3.connect(DB_PATH)
        subjects = pd.read_sql_query("SELECT id, subject_name FROM subjects ORDER BY subject_name", conn)
        conn.close()
        if subjects.empty:
            st.info("No subjects added yet.")
        else:
            st.dataframe(subjects, use_container_width=True)

        st.markdown("### 🗑️ Delete a Subject")
        if subjects.empty:
            st.info("No subjects to delete.")
        else:
            del_subj = st.selectbox("Select subject to delete", subjects["subject_name"].tolist(), key="del_subj_select")
            if st.button("🗑️ Delete Subject", key="del_subj_btn"):
                conn = sqlite3.connect(DB_PATH)
                conn.execute("DELETE FROM subjects WHERE subject_name=?", (del_subj,))
                conn.execute("DELETE FROM attendance WHERE subject=?", (del_subj,))
                conn.commit()
                conn.close()
                st.success(f"✅ Subject **{del_subj}** and all its attendance records deleted.")
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# 4. MARK ATTENDANCE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📷 Mark Attendance":
    st.title("📷 Mark Attendance")
    st.markdown("Real-time face recognition attendance marking")
    st.markdown("---")

    conn = sqlite3.connect(DB_PATH)
    subjects = conn.execute("SELECT subject_name FROM subjects ORDER BY subject_name").fetchall()
    conn.close()

    if not subjects:
        st.warning("⚠️ No subjects available. Please add subjects first.")
    else:
        subject_list = [s[0] for s in subjects]
        selected_subject = st.selectbox("📚 Select Subject", subject_list)

        col1, col2 = st.columns([2, 1])

        enc_data = load_encodings()
        if not enc_data["encodings"]:
            st.warning("⚠️ No students registered. Please register students first.")
        else:
            start_btn = col1.button("▶️ Start Attendance Camera")
            stop_btn  = col1.button("⏹️ Stop Camera")

            if start_btn:
                st.session_state["attendance_running"] = True
                st.session_state["recently_marked"] = {}

            if stop_btn:
                st.session_state["attendance_running"] = False

            if st.session_state.get("attendance_running", False):
                frame_placeholder = col1.empty()
                log_placeholder   = col2.empty()
                marked_log        = []

                recently_marked = st.session_state.get("recently_marked", {})
                cap = cv2.VideoCapture(0)

                if not cap.isOpened():
                    st.error("Cannot open webcam.")
                else:
                    run_limit = 60  # seconds
                    start = time.time()

                    while time.time() - start < run_limit:
                        if not st.session_state.get("attendance_running", False):
                            break

                        ret, frame = cap.read()
                        if not ret:
                            break

                        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        small = cv2.resize(rgb, (0, 0), fx=0.25, fy=0.25)
                        locs = face_recognition.face_locations(small)
                        encs = face_recognition.face_encodings(small, locs)

                        # Flip background first, draw boxes on mirrored coords
                        display = cv2.flip(rgb, 1)
                        w = display.shape[1]

                        for (top, right, bottom, left), enc in zip(locs, encs):
                            top *= 4; right *= 4; bottom *= 4; left *= 4
                            label = "Unknown"
                            color = (220, 50, 50)
                            roll_val = None

                            if enc_data["encodings"]:
                                matches = face_recognition.compare_faces(enc_data["encodings"], enc, tolerance=0.5)
                                face_distances = face_recognition.face_distance(enc_data["encodings"], enc)
                                best_idx = np.argmin(face_distances)

                                if matches[best_idx]:
                                    label = enc_data["names"][best_idx]
                                    roll_val = enc_data["rolls"][best_idx]
                                    color = (50, 220, 100)

                            # Mirror the x coordinates for drawing
                            m_left  = w - right
                            m_right = w - left
                            cv2.rectangle(display, (m_left, top), (m_right, bottom), color, 2)
                            cv2.putText(display, label, (m_left, top - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

                            # Mark attendance
                            if label != "Unknown" and roll_val:
                                key = f"{roll_val}_{selected_subject}"
                                now_ts = time.time()

                                if key not in recently_marked or (now_ts - recently_marked[key]) > 5:
                                    today = datetime.now().strftime("%Y-%m-%d")
                                    t_now = datetime.now().strftime("%H:%M:%S")

                                    conn = sqlite3.connect(DB_PATH)
                                    dup = conn.execute(
                                        "SELECT 1 FROM attendance WHERE roll=? AND subject=? AND date=?",
                                        (roll_val, selected_subject, today)
                                    ).fetchone()

                                    if not dup:
                                        conn.execute(
                                            "INSERT INTO attendance (name,roll,subject,date,time) VALUES (?,?,?,?,?)",
                                            (label, roll_val, selected_subject, today, t_now)
                                        )
                                        conn.commit()
                                        marked_log.append(f"✅ {label} ({roll_val})")

                                    conn.close()
                                    recently_marked[key] = now_ts

                        st.session_state["recently_marked"] = recently_marked
                        frame_placeholder.image(display, channels="RGB", use_column_width=True)

                        if marked_log:
                            log_placeholder.markdown("### 📋 Marked\n" + "\n".join(marked_log[-10:]))

                    cap.release()
                    st.info("Camera session ended.")

# ══════════════════════════════════════════════════════════════════════════════
# 5. REPORTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 Reports":
    st.title("📋 Attendance Reports")
    st.markdown("Subject-wise attendance analysis with percentage calculation")
    st.markdown("---")

    conn = sqlite3.connect(DB_PATH)
    subjects = conn.execute("SELECT subject_name FROM subjects ORDER BY subject_name").fetchall()
    conn.close()

    if not subjects:
        st.info("No subjects available. Please add subjects first.")
    else:
        subject_list = [s[0] for s in subjects]
        selected = st.selectbox("📚 Select Subject", subject_list)

        if st.button("📊 Generate Report"):
            conn = sqlite3.connect(DB_PATH)

            # Total unique class dates for subject
            total_classes_row = conn.execute(
                "SELECT COUNT(DISTINCT date) FROM attendance WHERE subject=?", (selected,)
            ).fetchone()
            total_classes = total_classes_row[0] if total_classes_row else 0

            # All students
            students_df = pd.read_sql_query("SELECT name, roll FROM students", conn)

            if students_df.empty:
                st.warning("No students registered yet.")
                conn.close()
            elif total_classes == 0:
                st.warning(f"No attendance records found for subject **{selected}**.")
                conn.close()
            else:
                # Attendance per student
                att_df = pd.read_sql_query(
                    "SELECT roll, COUNT(DISTINCT date) as classes_attended FROM attendance WHERE subject=? GROUP BY roll",
                    conn, params=(selected,)
                )
                conn.close()

                report = students_df.merge(att_df, on="roll", how="left")
                report["classes_attended"] = report["classes_attended"].fillna(0).astype(int)
                report["total_classes"] = total_classes
                report["percentage"] = (report["classes_attended"] / total_classes * 100).round(2)

                report = report.rename(columns={
                    "name": "Name",
                    "roll": "Roll Number",
                    "classes_attended": "Classes Attended",
                    "total_classes": "Total Classes",
                    "percentage": "Percentage (%)"
                })[["Name", "Roll Number", "Classes Attended", "Total Classes", "Percentage (%)"]]

                st.markdown(f"### 📊 Report: **{selected.upper()}** | Total Classes: **{total_classes}**")

                def color_percentage(val):
                    if val >= 75:
                        return "color: #00d4aa; font-weight: bold"
                    elif val >= 50:
                        return "color: #ffb432; font-weight: bold"
                    else:
                        return "color: #ff5050; font-weight: bold"

                styled = report.style.applymap(color_percentage, subset=["Percentage (%)"])
                st.dataframe(styled, use_container_width=True)

                avg = report["Percentage (%)"].mean()
                c1, c2, c3 = st.columns(3)
                c1.metric("📅 Total Classes", total_classes)
                c2.metric("👥 Students", len(report))
                c3.metric("📈 Avg Attendance", f"{avg:.1f}%")

# ══════════════════════════════════════════════════════════════════════════════
# 6. EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Export Excel":
    st.title("📤 Export Attendance to Excel")
    st.markdown("Download subject-wise attendance report as Excel file")
    st.markdown("---")

    conn = sqlite3.connect(DB_PATH)
    subjects = conn.execute("SELECT subject_name FROM subjects ORDER BY subject_name").fetchall()
    conn.close()

    if not subjects:
        st.info("No subjects available. Please add subjects first.")
    else:
        subject_list = [s[0] for s in subjects]
        selected = st.selectbox("📚 Select Subject to Export", subject_list)

        if st.button("📥 Generate Excel"):
            conn = sqlite3.connect(DB_PATH)
            total_classes_row = conn.execute(
                "SELECT COUNT(DISTINCT date) FROM attendance WHERE subject=?", (selected,)
            ).fetchone()
            total_classes = total_classes_row[0] if total_classes_row else 0

            students_df = pd.read_sql_query("SELECT name, roll FROM students", conn)

            if students_df.empty:
                st.warning("No students registered.")
                conn.close()
            else:
                att_df = pd.read_sql_query(
                    "SELECT roll, COUNT(DISTINCT date) as classes_attended FROM attendance WHERE subject=? GROUP BY roll",
                    conn, params=(selected,)
                )
                conn.close()

                report = students_df.merge(att_df, on="roll", how="left")
                report["classes_attended"] = report["classes_attended"].fillna(0).astype(int)
                report["total_classes"] = total_classes
                report["percentage"] = (report["classes_attended"] / max(total_classes, 1) * 100).round(2)
                report = report.rename(columns={
                    "name": "Name",
                    "roll": "Roll Number",
                    "classes_attended": "Classes Attended",
                    "total_classes": "Total Classes",
                    "percentage": "Percentage (%)"
                })[["Name", "Roll Number", "Classes Attended", "Total Classes", "Percentage (%)"]]

                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    report.to_excel(writer, index=False, sheet_name=selected[:31])
                    wb = writer.book
                    ws = writer.sheets[selected[:31]]

                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill("solid", fgColor="1a1a4a")
                    header_font = Font(bold=True, color="FFFFFF", size=12)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    for col in ws.columns:
                        max_len = max(len(str(cell.value or "")) for cell in col) + 4
                        ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

                output.seek(0)
                st.success("✅ Excel report generated!")
                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=output,
                    file_name=f"attendance_{selected}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ══════════════════════════════════════════════════════════════════════════════
# 7. RESET SYSTEM
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 Reset System":
    st.title("🔄 Reset System")
    st.markdown("⚠️ **Warning:** This will permanently delete ALL data.")
    st.markdown("---")

    st.error("This action will delete all students, attendance records, subjects, and face encodings. This cannot be undone.")

    col1, col2 = st.columns([1, 3])
    confirm = col1.checkbox("I understand the consequences")

    if confirm:
        if st.button("🗑️ RESET ENTIRE SYSTEM", type="primary"):
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM students")
            conn.execute("DELETE FROM attendance")
            conn.execute("DELETE FROM subjects")
            conn.execute("DELETE FROM sqlite_sequence WHERE name='students'")
            conn.execute("DELETE FROM sqlite_sequence WHERE name='attendance'")
            conn.execute("DELETE FROM sqlite_sequence WHERE name='subjects'")
            conn.commit()
            conn.close()

            if os.path.exists(ENCODINGS_FILE):
                os.remove(ENCODINGS_FILE)

            st.session_state.clear()
            st.success("✅ System has been completely reset. All data has been deleted.")
            st.balloons()
